import requests
import re
import os
import sys
from PyQt5.QtWidgets import QApplication, QMessageBox, QComboBox


from a_gui_from_ui import *

from datetime import datetime
from openpyxl import load_workbook


# запускаем гуй после конвертации ui в py
class Mywindow(QtWidgets.QMainWindow, Ui_MainWindow, QMessageBox, QComboBox):
    def __init__(self):
        super().__init__()
        self.setupUi(self)  # Инициализируем пользовательский интерфейс

        # добавляем функционал кнопок - ссылка на функции парсинг и сохр в эксель
        self.pushButton_Parse.clicked.connect(self.inn_check)
        self.pushButton_Record_to_Excel.clicked.connect(self.save_to_excel)

        # для выбора xlsm файла
        self.add_files_to_combo()

        self.excel_start()

    def inn_check(self):
        inn = self.lineEdit_INN_QT_INPUT.text()

        # проверка правильности ввода ИНН

        if len(inn) == 10 or len(inn) == 12 or len(inn) == 13 or len(inn) == 15:
            self.parse(inn)
        else:
            print('Ошибка в ИНН/ОГРН. Попробуйте еще раз')
            self.show_error_notification()

    def parse(self, inn):
        # сам парсинг
        url = 'https://egrul.nalog.ru'
        url_1 = 'https://egrul.nalog.ru/search-result/'

        r = requests.post(url, data={'query': inn})  # вставить инн в url в раздел дата
        r1 = requests.get(url_1 + r.json()['t'])
        r2 = r1.json()

        self.add_to_gui_parse_results(inn, r2)

    def add_to_gui_parse_results (self, inn, r2):

        # вот тут парсер
        if len(inn) == 10 or len(inn) == 13:
            nameOOO = str(r2['rows'][0]['c'])  # название контрагента ООО "КТото"
            nameOOO_here = re.search(r'(?<=").*?(?=")', nameOOO)  # найти текст между кавычек # print(nameOOO_first[0]) нужен индекс
            self.lineEdit_NAME_organization.setText(nameOOO_here[0])

            nameOOO_form = re.split(r'"', nameOOO)  # разделить строку на словарь по символу " # print(nameOOO_last[0])
            self.lineEdit_form.setText(nameOOO_form[0])

            adressOOO = str(r2['rows'][0]['a'])
            self.textEdit_adress.setText(adressOOO)

            seoOOO = str(r2['rows'][0]['g'])
            seoOOO_name = re.split(r': ', seoOOO)  # print(seoOOO_name [0]) и print(seoOOO_seo [1])
            self.lineEdit_seo_name.setText(seoOOO_name[1])

            seoOOO_director_position = str(seoOOO_name[0]).capitalize()
            self.lineEdit_seo_director_position.setText(seoOOO_director_position)

            # self добавляем чтобы можно было использовать из другого метода(другой функции) класса
            self.innOOO = str(r2['rows'][0]['i'])
            self.ogrnOOO = str(r2['rows'][0]['o'])
            self.kppOOO = str(r2['rows'][0]['p'])
        # для ИП
        elif len(inn) == 12 or len(inn) == 15:
            nameOOO = str.title(r2['rows'][0]['n'])  # str.title - первая буква заглавнаяу ФИО ИП
            self.lineEdit_NAME_organization.setText(nameOOO)
            self.kppOOO = str('')
            self.ogrnOOO = str(r2['rows'][0]['o'])
            self.innOOO = str(r2['rows'][0]['i'])
            self.lineEdit_form.setText('ИП')
        else:
            print("Ошибка в ИНН/ОГРН")

        #выбор xlsm файла
    def add_files_to_combo(self):
        current_directory = os.path.dirname(os.path.abspath(__file__))

        for filename in os.listdir(current_directory):
            if filename.endswith(".xlsm"):
                self.comboBox_xslm_choice.addItem(filename)
    def excel_start(self):
        self.selected_xlsm = self.comboBox_xslm_choice.currentText()
        self.choice_my_organization = re.search(r' (?P<found_text>.*?)\.', self.selected_xlsm)[1]  # найти текст между пробелом и точкой а [1] убирает пробел

        self.wb = load_workbook(self.selected_xlsm, read_only=False, keep_vba=True)  # после фн аргументы, чтобы можно было читать xlsm с макросами
        self.ws = self.wb['BD']  # имя листа
        self.last_record = (int(self.ws.max_row) + 1)  # найти номер незаполненной строки

        self.lineEdit_number_dogovor.setText(f'{self.last_record} {self.choice_my_organization}')
        self.lineEdit_date.setText(datetime.today().strftime('%d.%m.%Y'))
    def save_to_excel(self):

        # print(str('A') + str(last_record)) # номер ячейки
        # ячейка А
        dogovor_excel_formula = str('=CONCATENATE(' + str("S" + str(self.last_record))) + str('," Договор № ",E') + str(
            self.last_record) + str('," от ",TEXT(F') + str(self.last_record) + str(',"ДД.ММ.ГГ "),D') + str(self.last_record) + str(
            ')')
        self.ws[str('A') + str(self.last_record)] = dogovor_excel_formula

        # ячейки парсера забираем из введенного текста в форму
        self.ws[str('D') + str(self.last_record)] = self.lineEdit_NAME_organization.text()
        self.ws[str('J') + str(self.last_record)] = self.lineEdit_form.text()
        self.ws[str('L') + str(self.last_record)] = self.lineEdit_seo_director_position.text()
        self.ws[str('N') + str(self.last_record)] = self.lineEdit_seo_name.text()
        self.ws[str('V') + str(self.last_record)] = self.kppOOO
        self.ws[str('W') + str(self.last_record)] = self.textEdit_adress.toPlainText()
        self.ws[str('U') + str(self.last_record)] = self.ogrnOOO
        self.ws[str('T') + str(self.last_record)] = self.innOOO

        # остальные ячейки заполенные в qtDesigner
        self.ws[str('B') + str(self.last_record)] = self.comboBox_original.currentText()
        self.ws[str('E') + str(self.last_record)] = self.lineEdit_number_dogovor.text()
        self.ws[str('F') + str(self.last_record)] = self.lineEdit_date.text()
        self.ws[str('G') + str(self.last_record)] = self.lineEdit_name_dogovor.text()
        self.ws[str('R') + str(self.last_record)] = self.lineEdit_ustav.text()

        # ячейки формулы excel
        cell_k = '=IF(JXXX="ООО ","Общество с ограниченной ответственностью",IF(JXXX="АО ","Акционерное общество",IF(JXXX="НАО ","Непубличное кционерное общество",IF(JXXX="ПАО ","Публичное акционерное общество",IF(JXXX="ИП","Индивидуальный предприниматель",JXXX)))))'
        cell_k_change = re.sub(r'XXX', str(self.last_record), cell_k)
        self.ws[str('K') + str(self.last_record)] = cell_k_change

        cell_m = str('=IF(LYYY="Директор","директора",IF(LYYY="Генеральный директор","генерального директора",LYYY))')
        cell_m_change = re.sub(r'YYY', str(self.last_record), cell_m)
        self.ws[str('M') + str(self.last_record)] = str(cell_m_change)

        cell_o = '=LEFT(NUUU,SEARCH(" *",NUUU)-1)&" "&MID(NUUU,SEARCH(" *",NUUU)+1,1)&"."&MID(NUUU,SEARCH(" *",NUUU,SEARCH(" *",NUUU)+1)+1,1)&"."'
        cell_o_change = re.sub(r'UUU', str(self.last_record), cell_o)
        self.ws[str('O') + str(self.last_record)] = str(cell_o_change)

        cell_p = '=GenitiveCaseInCell1(NFFF)'
        cell_p_change = re.sub(r'FFF', str(self.last_record), cell_p)
        self.ws[str('P') + str(self.last_record)] = str(cell_p_change)

        cell_q = '=IF(RIGHT(NZZZ,1)="ч","его","ей")'
        cell_q_change = re.sub(r'ZZZ', str(self.last_record), cell_q)
        self.ws[str('Q') + str(self.last_record)] = str(cell_q_change)
        self.ws[str('R') + str(self.last_record)] = 'Устава'
        self.ws[str('S') + str(self.last_record)] = str(f'{self.last_record:05}')

        # Сохраняем файл
        self.wb.save(self.selected_xlsm)

        print("Текст сохранен в ячейке A1")
        print(self.innOOO)

    def show_error_notification(self):
        error_box = QMessageBox()  # Создаем окно уведомления
        error_box.setIcon(QMessageBox.Critical)  # Устанавливаем иконку ошибки
        error_box.setWindowTitle("Ошибка")  # Устанавливаем заголовок окна
        error_box.setText("Ошибка в номере ИНН/ОГРН. Исправьте")  # Устанавливаем текст ошибки
        error_box.exec_()  # Показываем окно уведомления и блокируем основное окно приложения


app = QApplication(sys.argv)
window = Mywindow()
window.show()
sys.exit(app.exec_())