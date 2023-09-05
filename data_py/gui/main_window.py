import requests
import re
import os
import sys
import configparser

import data_py.gui.content.qrcode_rc

from data_py.gui.a_gui_from_ui import *
from data_py.gui.b_gui_about import Ui_Form_about

from PyQt5.QtWidgets import QApplication, QMessageBox, QComboBox, QAction
from PyQt5.QtCore import QTimer, QUrl
from PyQt5.QtGui import QDesktopServices
from datetime import datetime
from openpyxl import load_workbook


class AboutWindow(QtWidgets.QMainWindow, Ui_Form_about):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

class Mywindow(QtWidgets.QMainWindow, Ui_MainWindow, QMessageBox, QComboBox, QAction):
    def __init__(self):
        super().__init__()
        self.setupUi(self)  # Инициализируем пользовательский интерфейс

        # добавляем функционал кнопок
        self.pushButton_Parse.clicked.connect(self.inn_check)
        self.pushButton_Record_to_Excel.clicked.connect(self.save_to_excel)
        self.pushButton_cancel_record.clicked.connect(self.cancel_record)
        self.pushButton_donation.clicked.connect(self.openAboutWindow)
        self.checkBox_no_parser.stateChanged.connect(self.no_parser)

        # для выбора xlsm файла
        self.add_files_to_combo()
        self.excel_start()
        self.comboBox_xslm_choice.currentIndexChanged.connect(self.excel_start)

        # menu bar
        # menu bar добавить фукнционал для  открыть excel конфиг и др вкладки
        self.action_open_excel.triggered.connect(lambda: os.startfile(self.selected_xlsm))
        self.action_settings.triggered.connect(lambda: os.startfile("data_py\config.ini"))

        self.action_clear_inn.triggered.connect(self.second_parse_clear_results)
        self.action_clear_default.triggered.connect(self.clear_default)
        self.action_clear_all.triggered.connect(self.clear_all)
        self.config_settings()


        self.action_readme.triggered.connect(lambda: os.startfile("readme.txt"))
        self.action_info_author.triggered.connect(self.openAboutWindow)


        self.actionTelegram.triggered.connect(lambda: QDesktopServices.openUrl(QUrl("https://t.me/pixelpravo")))
        self.actionInstagram.triggered.connect(lambda: QDesktopServices.openUrl(QUrl("https://instagram.com/zimzangle?utm_source=qr&igshid=MzNlNGNkZWQ4Mg==")))
        self.actionPixelpravo_ru.triggered.connect(lambda: QDesktopServices.openUrl(QUrl("https://www.pixelpravo.ru")))


    def inn_check(self):
        try:
            inn = self.lineEdit_INN_QT_INPUT.text()

            # проверка правильности ввода ИНН

            if len(inn) == 10 or len(inn) == 12 or len(inn) == 13 or len(inn) == 15:
                self.parse(inn)
            else:
                message_error = ('ИНН/ОГРН введен неверно. Исправьте')
                self.show_error_notification(message_error)
        except:
            message_error = ('Ошибка. Возможно, этот ИНН (АНО и другие) не поддерживается')
            self.show_error_notification(message_error)

    def parse(self, inn):
        # сам парсинг
        url = 'https://egrul.nalog.ru'
        url_1 = 'https://egrul.nalog.ru/search-result/'

        r = requests.post(url, data={'query': inn})  # вставить инн в url в раздел дата
        r1 = requests.get(url_1 + r.json()['t'])
        r2 = r1.json()

        self.add_to_gui_parse_results(inn, r2)

    def add_to_gui_parse_results (self, inn, r2):

        #убрать старые результаты
        self.second_parse_clear_results()
        # разбив результата парсинга на элементы
        if len(inn) == 10 or len(inn) == 13:
            nameOOO = str(r2['rows'][0]['c'])  # название контрагента ООО "КТото"
            nameOOO_here = re.search(r'(?<=").*?(?=")', nameOOO)  # найти текст между кавычек # print(nameOOO_first[0]) нужен индекс
            self.lineEdit_NAME_organization.setText(nameOOO_here[0])

            nameOOO_form = re.split(r'"', nameOOO)  # разделить строку на словарь по символу " # print(nameOOO_last[0])
            self.lineEdit_form.setText(nameOOO_form[0])

            adressOOO = str(r2['rows'][0]['a'])
            self.textEdit_adress.setText(adressOOO)

            seoOOO = str(r2['rows'][0]['g'])
            seoOOO_name = re.split(r': ', seoOOO)  # print(seoOOO_name [0]) и print(seoOOO_seo [1])
            self.lineEdit_seo_name.setText(seoOOO_name[1])

            seoOOO_director_position = str(seoOOO_name[0]).capitalize()
            self.lineEdit_seo_director_position.setText(seoOOO_director_position)

            # self добавляем чтобы можно было использовать из другого метода(другой функции) класса
            self.innOOO = str(r2['rows'][0]['i'])
            self.ogrnOOO = str(r2['rows'][0]['o'])
            self.kppOOO = str(r2['rows'][0]['p'])

            # переключение на 1 вкладку
            self.tabWidget.setCurrentIndex(0)
        # для ИП
        elif len(inn) == 12 or len(inn) == 15:

            # переключение на 2 вкладку для ИП
            self.tabWidget.setCurrentIndex(1)

            nameOOO = str.title(r2['rows'][0]['n'])  # str.title - первая буква заглавнаяу ФИО ИП
            self.lineEdit_NAME_organization_2.setText(nameOOO)
            self.kppOOO = str('')
            self.ogrnOOO = str(r2['rows'][0]['o'])
            self.innOOO = str(r2['rows'][0]['i'])
            self.lineEdit_form.setText('ИП')
        else:
            print("Ошибка в ИНН/ОГРН")

        #выбор xlsm файла добавление из директории назаваний

    def add_files_to_combo(self):

        # os.path.dirname трижды чтобы искал не в Project2\data_py\gui а в Project2\ было так current_directory = os.path.dirname(os.path.abspath(__file__))
        current_directory = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

        for filename in os.listdir(current_directory):
            if filename.endswith(".xlsm"):
                self.comboBox_xslm_choice.addItem(filename)
    def excel_start(self):
        self.selected_xlsm = self.comboBox_xslm_choice.currentText()
        self.choice_my_organization = re.search(r' (?P<found_text>.*?)\.', self.selected_xlsm)[1]  # найти текст между пробелом и точкой а [1] убирает пробел

        self.wb = load_workbook(self.selected_xlsm, read_only=False, keep_vba=True)  # после фн аргументы, чтобы можно было читать xlsm с макросами
        self.ws = self.wb['BD']  # имя листа
        self.last_record = (int(self.ws.max_row) + 1)  # найти номер незаполненной строки

        self.lineEdit_number_docs.setText(f'{self.last_record} {self.choice_my_organization}')
        self.lineEdit_date.setText(datetime.today().strftime('%d.%m.%Y'))
    def save_to_excel(self):
        try:
            # print(str('A') + str(last_record)) # номер ячейки
            # ячейка А
            cell_a = '=CONCATENATE(SXXX, " ",GXXX," № ", EXXX, " ", IXXX, " от ", TEXT(FXXX, "ДД.ММ.ГГ "), DXXX)'
            cell_a_change = re.sub(r'XXX', str(self.last_record), cell_a)
            self.ws[f'{"A"}{self.last_record}'] = cell_a_change

            ## ячейки парсера забираем из введенного текста в форму

            #if esle разграничение выбранной вкладки ЮЛ и ИП
            if self.tabWidget.currentIndex() == 0:
                self.ws[f'{"D"}{self.last_record}'] = self.lineEdit_NAME_organization.text()
                self.ws[f'{"W"}{self.last_record}'] = self.textEdit_adress.toPlainText()
            else:
                self.ws[f'{"D"}{self.last_record}'] = self.lineEdit_NAME_organization_2.text()
                self.ws[f'{"W"}{self.last_record}'] = self.textEdit_adress_2.toPlainText()
                self.ws[f'{"AC"}{self.last_record}'] = self.textEdit_pasport.toPlainText()



            self.ws[f'{"J"}{self.last_record}'] = self.lineEdit_form.text()
            self.ws[f'{"L"}{self.last_record}'] = self.lineEdit_seo_director_position.text()
            self.ws[f'{"N"}{self.last_record}'] = self.lineEdit_seo_name.text()
            self.ws[f'{"V"}{self.last_record}'] = self.kppOOO

            self.ws[f'{"U"}{self.last_record}'] = self.ogrnOOO
            self.ws[f'{"T"}{self.last_record}'] = self.innOOO

            # остальные ячейки заполенные в qtDesigner
            self.ws[f'{"B"}{self.last_record}'] = self.comboBox_docfate.currentText()
            self.ws[f'{"C"}{self.last_record}'] = self.lineEdit_time_to_pay.text()
            self.ws[f'{"E"}{self.last_record}'] = self.lineEdit_number_docs.text()
            self.ws[f'{"F"}{self.last_record}'] = self.lineEdit_date.text()
            self.ws[f'{"G"}{self.last_record}'] = self.comboBox_type_doc.currentText()
            self.ws[f'{"H"}{self.last_record}'] = self.comboBox_name_doc.currentText()
            self.ws[f'{"AA"}{self.last_record}'] = self.lineEdit_number_email.text()
            self.ws[f'{"AB"}{self.last_record}'] = self.lineEdit_debt_sum.text()


            # ячейки неизменяемые формулы excel
            cell_k = '=IF(JXXX="ООО ","Общество с ограниченной ответственностью",IF(JXXX="АО ","Акционерное общество",IF(JXXX="НАО ","Непубличное кционерное общество",IF(JXXX="ПАО ","Публичное акционерное общество",IF(JXXX="ИП","Индивидуальный предприниматель",JXXX)))))'
            cell_k_change = re.sub(r'XXX', str(self.last_record), cell_k)
            self.ws[f'{"K"}{self.last_record}'] = cell_k_change

            cell_m = '=IF(LXXX="Директор","директора",IF(LXXX="Генеральный директор","генерального директора",LXXX))'
            cell_m_change = re.sub(r'XXX', str(self.last_record), cell_m)
            self.ws[f'{"M"}{self.last_record}'] = cell_m_change

            cell_o = '=LEFT(NXXX,SEARCH(" *",NXXX)-1)&" "&MID(NXXX,SEARCH(" *",NXXX)+1,1)&"."&MID(NXXX,SEARCH(" *",NXXX,SEARCH(" *",NXXX)+1)+1,1)&"."'
            cell_o_change = re.sub(r'XXX', str(self.last_record), cell_o)
            self.ws[f'{"O"}{self.last_record}'] = cell_o_change

            cell_p = '=GenitiveCaseInCell1(NXXX)'
            cell_p_change = re.sub(r'XXX', str(self.last_record), cell_p)
            self.ws[f'{"P"}{self.last_record}'] = cell_p_change

            cell_q = '=IF(RIGHT(NXXX,1)="ч","его","ей")'
            cell_q_change = re.sub(r'XXX', str(self.last_record), cell_q)
            self.ws[f'{"Q"}{self.last_record}'] = cell_q_change
            self.ws[f'{"S"}{self.last_record}'] = f'{self.last_record:05}'


            # для писем и претензий
            cell_x = '=IF(LXXX="Директор","Директору",IF(LXXX="Генеральный директор","Генеральному директору",LXXX))'
            cell_x_change = re.sub(r'XXX', str(self.last_record), cell_x)
            self.ws[f'{"X"}{self.last_record}'] = cell_x_change
            self.ws[f'{"Y"}{self.last_record}'] = f'{"=DativeCase(N"}{self.last_record}{")"}'
            cell_z = '=LEFT(YXXX,SEARCH(" *",YXXX)-1)&" "&MID(YXXX,SEARCH(" *",YXXX)+1,1)&"."&MID(YXXX,SEARCH(" *",YXXX,SEARCH(" *",YXXX)+1)+1,1)&"."'
            cell_z_change = re.sub(r'XXX', str(self.last_record), cell_z)
            self.ws[f'{"Z"}{self.last_record}'] = cell_z_change

            # Сохраняем файл
            self.wb.save(self.selected_xlsm)

            message = f'Успешно записано в excel в {self.choice_my_organization}'
            self.message_to_user(message)

            # обновляем номер договора
            self.excel_start()

            # кнопка отмены
            self.pushButton_cancel_record.setEnabled(True)

        except:
            message_error = '-Закрой excel файл с реестром, который сейчас открыт\nили\n-не была нажата кнопка "Подтвердить вввод ИНН"\nили\n-не поставлена галочка на ввод без ИНН'
            self.show_error_notification(message_error)

    def show_error_notification(self, message_error):
        error_box = QMessageBox()  # Создаем окно уведомления
        error_box.setIcon(QMessageBox.Critical)  # Устанавливаем иконку ошибки
        error_box.setWindowTitle("Ошибка")  # Устанавливаем заголовок окна
        error_box.setText(message_error)  # Устанавливаем текст ошибки
        error_box.exec_()  # Показываем окно уведомления и блокируем основное окно приложения

    def second_parse_clear_results(self):
        self.lineEdit_NAME_organization.clear()
        self.textEdit_adress.clear()
        self.lineEdit_seo_director_position.clear()
        self.lineEdit_form.clear()
        self.lineEdit_seo_name.clear()
        # у ИП
        self.lineEdit_NAME_organization_2.clear()
        self.textEdit_adress_2.clear()
        self.textEdit_pasport.clear()

    def clear_all(self):
        self.second_parse_clear_results()

        self.lineEdit_INN_QT_INPUT.clear()
        self.lineEdit_date.clear()
        self.lineEdit_number_docs.clear()
        self.lineEdit_time_to_pay.clear()
        self.lineEdit_number_email.clear()
        self.lineEdit_debt_sum.clear()
        self.comboBox_type_doc.clear()
        self.comboBox_name_doc.clear()
        self.comboBox_docfate.clear()

        # возврат значений  combobox из конфига и из excel
    def clear_default(self):
        self.clear_all()

        self.config_settings()
        self.excel_start()

    def no_parser(self, state):
        if state == 2:
            self.lineEdit_INN_QT_INPUT.setEnabled(False)
            self.kppOOO = ""
            self.ogrnOOO = ""
            self.innOOO = ""

        else:
            self.lineEdit_INN_QT_INPUT.setEnabled(True)
            del self.kppOOO
            del self.ogrnOOO
            del self.innOOO


    def message_to_user(self, message):
        self.lineEdit_result.setText(message)
        self.timer = QTimer(self)
        self.timer.start(15000)  # 15000 миллисекунд (15 секунд)
        self.timer.timeout.connect(lambda: self.lineEdit_result.clear())

    def config_settings(self):
        # Чтение значений из конфигурационного файла
        config = configparser.ConfigParser()
        config.read('data_py\config.ini', encoding='utf-8')
        combo_values_doctype = config['Doc Type']
        combo_values_docname = config['Doc Name']
        combo_values_docfate = config['Doc Fate']

        # Добавление значений в QComboBox
        self.comboBox_type_doc.addItems(combo_values_doctype.values())
        self.comboBox_name_doc.addItems(combo_values_docname.values())
        self.comboBox_docfate.addItems(combo_values_docfate.values())

    def cancel_record(self):
        self.ws.delete_rows(self.ws.max_row, amount=1)
        self.wb.save(self.selected_xlsm)
        self.excel_start()
        self.pushButton_cancel_record.setEnabled(False)
        message = f'Запись удалена из {self.choice_my_organization}'
        self.message_to_user(message)

    def openAboutWindow(self):
        self.about_window = AboutWindow()
        self.about_window.show()


app = QApplication(sys.argv)
window = Mywindow()
window.show()
sys.exit(app.exec_())